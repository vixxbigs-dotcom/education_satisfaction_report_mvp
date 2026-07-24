from __future__ import annotations

import io
import math
import re
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from .models import ChoiceQuestion, ObjectiveQuestion, ReportData, SubjectiveQuestion
from .text_utils import strip_leading_question_numbers


DEFAULT_SCALE = ["매우 그렇다", "그렇다", "보통이다", "그렇지 않다", "매우 그렇지 않다"]

TEXT_SCORE_MAP = {
    "매우그렇다": 5,
    "그렇다": 4,
    "보통이다": 3,
    "보통": 3,
    "그렇지않다": 2,
    "매우그렇지않다": 1,
    "매우만족": 5,
    "만족": 4,
    "불만족": 2,
    "매우불만족": 1,
    "매우우수": 5,
    "우수": 4,
    "미흡": 2,
    "매우미흡": 1,
    "매우도움": 5,
    "도움": 4,
    "도움안됨": 2,
    "전혀도움안됨": 1,
}

QUESTION_HINTS = (
    "만족", "추천", "적절", "난이도", "강의", "교육", "좋았", "아쉬", "개선",
    "운영", "도움", "평가", "의견", "작성", "전반적", "전문", "과정", "무엇",
    "어느", "선택", "적용", "이유", "제안", "충분", "원활", "명확", "경험"
)

METADATA_HEADERS = (
    "타임스탬프", "timestamp", "응답자", "성함", "이름", "소속", "부서", "직급", "조 구분",
    "조구분", "이메일", "email", "연락처", "사번", "순번"
)

TITLE_SUFFIXES = (
    "교수", "강사", "코치", "대표", "노무사", "변호사", "박사", "컨설턴트",
    "연구위원", "위원", "전문가", "선생님"
)

NO_OPINION_CORES = {
    "", "없", "없음", "없습니다", "없어요", "없다", "없었음", "없었습니다",
    "해당없음", "해당사항없음", "특별히없음", "특별히없었음", "딱히없음",
    "무", "무응답", "na", "n/a", "-", ".", "x"
}

MULTI_SEPARATORS = re.compile(r"\s*(?:;|\n|\r|\|)\s*|\s*,\s*")


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _norm(value: Any) -> str:
    return re.sub(r"[\s\u00a0]+", "", str(value or "")).strip().lower()


def _is_question(text: Any) -> bool:
    t = _clean_text(text)
    if len(t) < 4:
        return False
    if "?" in t or "습니까" in t or "작성해" in t or "무엇" in t:
        return True
    return any(k in t for k in QUESTION_HINTS)


def _is_metadata_header(text: str) -> bool:
    n = _norm(text)
    return any(_norm(token) in n for token in METADATA_HEADERS) and not _is_question(text)


def _to_numeric_score(value: Any) -> Optional[int]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        numeric = float(value)
        if abs(numeric - round(numeric)) > 1e-7:
            return None
        iv = int(round(numeric))
        return iv if 0 <= iv <= 10 else None
    text = _norm(value)
    if not text:
        return None
    match = re.fullmatch(r"(10|[0-9])(?:점)?", text)
    if match:
        return int(match.group(1))
    return TEXT_SCORE_MAP.get(text)


def _infer_scale(values: Sequence[Any], question: str = "") -> Optional[Tuple[int, int, List[int]]]:
    nonempty = [v for v in values if _clean_text(v)]
    if not nonempty:
        return None
    mapped = [_to_numeric_score(v) for v in nonempty]
    scores = [v for v in mapped if v is not None]
    hit_ratio = len(scores) / len(nonempty)
    if len(scores) < 2 or hit_ratio < 0.72:
        return None

    observed_min = min(scores)
    observed_max = max(scores)
    q_norm = _norm(question)

    # The response distribution is the primary signal. Header hints only resolve
    # incomplete samples where the top point was not selected by anyone.
    if observed_min == 0 or observed_max > 7 or "0~10" in q_norm or "0-10" in q_norm or "nps" in q_norm:
        scale_min, scale_max = 0, 10
    elif observed_max > 5:
        scale_min, scale_max = 1, 7
    elif observed_max <= 5:
        scale_min, scale_max = 1, 5
    else:
        return None

    if not all(scale_min <= score <= scale_max for score in scores):
        return None
    return scale_min, scale_max, scores


def _question_number(text: str, fallback: int) -> int:
    match = re.match(r"\s*(\d+)\s*[.)]?", text)
    return int(match.group(1)) if match else fallback


def _extract_module(text: str) -> str:
    match = re.search(r"\[\s*([^\]]+?)\s*\]", text)
    if match:
        return _clean_text(match.group(1))
    match = re.search(r"\(\s*([^()]{2,25})\s*\)", text)
    return _clean_text(match.group(1)) if match else ""


def _normalize_instructor_name(candidate: str) -> str:
    """Keep the display title because the PPT should show e.g. '설상훈 교수'."""
    candidate = _clean_text(candidate)
    candidate = re.sub(r"^.*?\]\s*", "", candidate)
    candidate = re.sub(r"^\d+\s*[.)]?\s*", "", candidate)
    words = candidate.split()
    if len(words) > 4:
        candidate = " ".join(words[-3:])
    return candidate


def _extract_instructor_slot(text: str) -> int:
    match = re.search(r"(?:\{|\[)?강사\s*(10|[1-9])(?:\}|\])?", _clean_text(text))
    return int(match.group(1)) if match else 0


def _extract_instructor(text: str) -> str:
    suffix = "|".join(map(re.escape, TITLE_SUFFIXES))
    # Name + title followed by possessive/subject marker is the strongest case.
    matches = re.findall(
        rf"([가-힣A-Za-z][가-힣A-Za-z·.\- ]{{1,20}}?\s*(?:{suffix}))(?=의|는|은|께서|\s|$)",
        text,
    )
    if matches:
        normalized = _normalize_instructor_name(matches[-1])
        # Generic words such as "강사" or "퍼실리테이터(강사)" are not names.
        if normalized and normalized not in {"강사", "퍼실리테이터", "담당", "외부"}:
            return normalized

    # Frequent Google Forms phrasing: "윤영철 강사의 ..."
    match = re.search(r"([가-힣A-Za-z]{2,12}\s*(?:강사|교수|코치))\s*의", text)
    if match:
        name = _clean_text(match.group(1))
        if name.split()[0] not in {"해당", "담당", "외부", "내부"}:
            return name
    return ""


def _instructor_metric(text: str) -> str:
    t = _clean_text(text)
    if any(k in t for k in ("전문", "지식", "전문성")):
        return "전문성"
    if any(k in t for k in ("이해하기 쉽게", "설명", "전달")):
        return "설명력"
    if any(k in t for k in ("진행", "강의 진행")):
        return "강의 진행"
    if any(k in t for k in ("질문", "가이드", "퍼실리테이")):
        return "퍼실리테이션"
    if any(k in t for k in ("만족", "전반")):
        return "전반 만족"
    return "기타"


def _classify_scale_question(text: str) -> Tuple[str, str]:
    t = _clean_text(text)
    instructor = _extract_instructor(t)
    module = _extract_module(t)
    if instructor:
        return "강사 만족도", "instructor"
    if "추천" in t or "nps" in t.lower():
        return "추천도", "recommendation"
    if "일정" in t or "시간" in t:
        return "교육 일정", "schedule"
    if "난이도" in t:
        return "교육 난이도", "difficulty"
    if any(k in t for k in ("장소", "사전 안내", "과정 운영", "교육 운영", "운영 만족", "운영 전반")):
        return "교육 운영", "operation"
    if "강사" in t or "퍼실리테이터" in t:
        return "강사 만족도", "instructor_generic"
    if "종합" in t and "만족" in t:
        return "종합 만족도", "overall"
    if "전반" in t and "만족" in t:
        return "전반적 만족도", "overall"
    if module or "과정에 만족" in t:
        return module or "과정 만족도", "course"
    return "만족도", "other"


def _classify_subjective(text: str) -> Tuple[str, str]:
    t = _clean_text(text)
    if any(k in t for k in ("좋았", "장점", "인상", "유익", "우수 요소", "만족스러운")):
        return "좋았던 점", "subjective_good"
    if any(k in t for k in ("아쉬", "개선", "보완", "불편", "단점", "건의", "바라는")):
        return "아쉬웠던 점", "subjective_bad"
    if any(k in t for k in ("적용", "현업")):
        return "현업 적용", "subjective_apply"
    return "주관식", "subjective_other"


def _choice_tokens(value: Any) -> List[str]:
    text = _clean_text(value)
    if not text:
        return []
    return [token.strip() for token in MULTI_SEPARATORS.split(text) if token.strip()]


def _detect_choice(values: Sequence[Any], question: str) -> Optional[Tuple[str, List[str], int]]:
    texts = [_clean_text(v) for v in values if _clean_text(v)]
    if len(texts) < 2:
        return None

    tokenized = [_choice_tokens(v) for v in texts]
    has_multi = any(len(tokens) > 1 for tokens in tokenized)
    all_tokens = [token for tokens in tokenized for token in tokens]
    token_counts = Counter(all_tokens)
    unique_tokens = list(token_counts)
    q_norm = _norm(question)

    # Multiple-choice requires repeated short option-like tokens. This avoids
    # treating free-text sentences containing commas as checkbox responses.
    if has_multi:
        repeated = sum(1 for count in token_counts.values() if count >= 2)
        short_ratio = sum(len(token) <= 35 for token in all_tokens) / max(1, len(all_tokens))
        if (
            "복수" in q_norm
            or "중복" in q_norm
            or (len(unique_tokens) <= max(12, len(texts)) and repeated >= 2 and len(unique_tokens) <= 12 and short_ratio >= 0.82)
        ):
            return "multiple", unique_tokens, len(texts)

    counts = Counter(texts)
    unique_count = len(counts)
    repeated_count = sum(1 for count in counts.values() if count >= 2)
    avg_length = sum(len(text) for text in texts) / len(texts)
    # Single-choice values repeat across respondents and form a limited set.
    if (
        "단일선택" in q_norm
        or "단일 선택" in question
        or "선택" in question
        or (
            2 <= unique_count <= min(12, max(3, math.floor(len(texts) * 0.5)))
            and unique_count / len(texts) <= 0.5
            and repeated_count >= 1
            and avg_length <= 45
        )
    ):
        return "single", list(counts.keys()), len(texts)
    return None


def _build_objective(question: str, values: Sequence[Any], fallback_no: int, qid: str, scale_info) -> ObjectiveQuestion:
    source_question = _clean_text(question)
    cleaned_question = strip_leading_question_numbers(source_question)
    scale_min, scale_max, scores = scale_info
    scale_values = list(range(scale_max, scale_min - 1, -1))
    counts = [scores.count(score) for score in scale_values]
    average = round(sum(scores) / len(scores), 2) if scores else 0.0
    section, category = _classify_scale_question(cleaned_question)
    labels = (
        list(DEFAULT_SCALE)
        if scale_min == 1 and scale_max == 5
        else [f"{value}점" for value in scale_values]
    )
    instructor = _extract_instructor(cleaned_question)
    return ObjectiveQuestion(
        id=qid,
        number=_question_number(source_question, fallback_no),
        section_label=section,
        question=cleaned_question,
        counts=counts,
        scale_labels=labels,
        average=average,
        valid_responses=len(scores),
        category=category,
        instructor_name=instructor,
        course_module=_extract_module(cleaned_question),
        scale_min=scale_min,
        scale_max=scale_max,
        scale_values=scale_values,
        scale_type="numeric" if all(isinstance(v, (int, float)) for v in values if _clean_text(v)) else "likert",
        instructor_metric=_instructor_metric(cleaned_question) if (instructor or _extract_instructor_slot(cleaned_question)) else "",
        summary_label="",
        instructor_slot=_extract_instructor_slot(cleaned_question),
    )


def _build_choice(question: str, values: Sequence[Any], fallback_no: int, qid: str, detection) -> ChoiceQuestion:
    source_question = _clean_text(question)
    cleaned_question = strip_leading_question_numbers(source_question)
    selection_type, options, valid_responses = detection
    if selection_type == "multiple":
        counter = Counter(token for value in values for token in _choice_tokens(value))
    else:
        counter = Counter(_clean_text(v) for v in values if _clean_text(v))
    ordered_options = sorted(counter.keys(), key=lambda option: (-counter[option], options.index(option) if option in options else 9999))
    counts = [counter[option] for option in ordered_options]
    percentages = [round(count / valid_responses * 100, 1) if valid_responses else 0.0 for count in counts]
    return ChoiceQuestion(
        id=qid,
        number=_question_number(source_question, fallback_no),
        section_label="선택형 문항",
        question=cleaned_question,
        options=ordered_options,
        counts=counts,
        percentages=percentages,
        selection_type=selection_type,
        valid_responses=valid_responses,
        category=f"choice_{selection_type}",
    )


def _build_subjective(question: str, values: Sequence[Any], fallback_no: int, qid: str) -> SubjectiveQuestion:
    source_question = _clean_text(question)
    cleaned_question = strip_leading_question_numbers(source_question)
    answers = [_clean_text(v) for v in values if _clean_text(v)]
    section, category = _classify_subjective(cleaned_question)
    return SubjectiveQuestion(
        id=qid,
        number=_question_number(source_question, fallback_no),
        section_label=section,
        question=cleaned_question,
        answers=answers,
        category=category,
    )


def _classify_and_build(question: str, values: Sequence[Any], fallback_no: int, qid: str):
    _, subjective_category = _classify_subjective(question)
    if subjective_category != "subjective_other" or any(token in _clean_text(question) for token in ("작성해", "자유롭게", "의견을")):
        return "subjective", _build_subjective(question, values, fallback_no, qid)
    scale_info = _infer_scale(values, question)
    if scale_info:
        return "objective", _build_objective(question, values, fallback_no, qid, scale_info)
    choice_info = _detect_choice(values, question)
    if choice_info:
        return "choice", _build_choice(question, values, fallback_no, qid, choice_info)
    return "subjective", _build_subjective(question, values, fallback_no, qid)


def _sheet_matrix(ws) -> List[List[Any]]:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    return [list(row) for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True)]


def _orientation_score(matrix: List[List[Any]]) -> Tuple[int, int]:
    if not matrix:
        return 0, 0
    first_row = matrix[0]
    first_col = [row[0] if row else None for row in matrix]
    row_score = sum(_is_question(v) for v in first_row)
    col_score = sum(_is_question(v) for v in first_col)
    return row_score, col_score


def _parse_column_oriented(matrix: List[List[Any]], sheet_name: str):
    headers = matrix[0] if matrix else []
    rows = matrix[1:]
    objective: List[ObjectiveQuestion] = []
    choices: List[ChoiceQuestion] = []
    subjective: List[SubjectiveQuestion] = []
    response_count = sum(1 for row in rows if any(_clean_text(v) for v in row))
    qidx = 1
    for ci, header in enumerate(headers):
        question = _clean_text(header)
        if not question or _is_metadata_header(question):
            continue
        if not _is_question(question):
            continue
        values = [row[ci] if ci < len(row) else None for row in rows]
        if not any(_clean_text(value) for value in values):
            continue
        kind, item = _classify_and_build(question, values, qidx, f"{sheet_name}-c{ci+1}")
        if kind == "objective":
            objective.append(item)
        elif kind == "choice":
            choices.append(item)
        else:
            subjective.append(item)
        qidx += 1
    return objective, choices, subjective, response_count


def _parse_row_oriented(matrix: List[List[Any]], sheet_name: str):
    objective: List[ObjectiveQuestion] = []
    choices: List[ChoiceQuestion] = []
    subjective: List[SubjectiveQuestion] = []
    response_count = 0
    qidx = 1
    for ri, row in enumerate(matrix):
        if not row:
            continue
        question = _clean_text(row[0])
        if not question or _is_metadata_header(question) or not _is_question(question):
            continue
        values = list(row[1:])
        if not any(_clean_text(value) for value in values):
            continue
        response_count = max(response_count, sum(1 for v in values if _clean_text(v)))
        kind, item = _classify_and_build(question, values, qidx, f"{sheet_name}-r{ri+1}")
        if kind == "objective":
            objective.append(item)
        elif kind == "choice":
            choices.append(item)
        else:
            subjective.append(item)
        qidx += 1
    return objective, choices, subjective, response_count


def _sheet_candidate(matrix: List[List[Any]], sheet_name: str) -> Optional[Dict[str, Any]]:
    row_score, col_score = _orientation_score(matrix)
    if max(row_score, col_score) == 0:
        return None
    if col_score > row_score:
        objective, choices, subjective, count = _parse_row_oriented(matrix, sheet_name)
        orientation = "문항이 행 / 응답자가 열"
    else:
        objective, choices, subjective, count = _parse_column_oriented(matrix, sheet_name)
        orientation = "문항이 열 / 응답자가 행"
    quality = len(objective) * 6 + len(choices) * 4 + len(subjective) * 2 + min(count, 30)
    if any(token in sheet_name.lower() for token in ("응답", "response", "로우")):
        quality += 80
    return {
        "sheet_name": sheet_name,
        "orientation": orientation,
        "objective": objective,
        "choices": choices,
        "subjective": subjective,
        "response_count": count,
        "quality": quality,
    }


def _all_cell_texts(workbook) -> List[str]:
    texts: List[str] = []
    for ws in workbook.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                text = _clean_text(value)
                if text:
                    texts.append(text)
    return texts


def _extract_metadata(texts: Iterable[str], filename: str) -> Dict[str, str]:
    joined = "\n".join(texts)
    metadata: Dict[str, str] = {"company_name": "", "course_name": "", "course_round": "", "schedule": ""}

    for pattern in (r"교육과정\s*[:：]\s*([^\n]+)", r"과정명\s*[:：]\s*([^\n]+)"):
        match = re.search(pattern, joined)
        if match:
            metadata["course_name"] = re.sub(r"^[■●\-]\s*", "", _clean_text(match.group(1)))
            break

    for pattern in (
        r"교육일시\s*[:：]\s*([^\n]+)",
        r"교육일정\s*[:：]\s*([^\n]+)",
        r"(20\d{2}[./-]\d{1,2}[./-]\d{1,2}\.?\s*[~～-]\s*20?\d{0,2}[./-]?\d{1,2}[./-]\d{1,2}\.?)",
    ):
        match = re.search(pattern, joined)
        if match:
            metadata["schedule"] = _clean_text(match.group(1))
            break

    clean_filename = Path(filename).stem
    clean_filename = re.sub(r"\([^)]*로우데이터[^)]*\)", "", clean_filename, flags=re.I)
    clean_filename = re.sub(r"_?결과본$", "", clean_filename)
    clean_filename = re.sub(r"교육 만족도 (설문|조사).*$", "", clean_filename)
    clean_filename = _clean_text(clean_filename)
    if not metadata["course_name"]:
        metadata["course_name"] = clean_filename

    for pattern in (r"^\(([^)]+)\)", r"^([A-Za-z][A-Za-z0-9 &.-]{1,25})\s+", r"^(한국[가-힣A-Za-z0-9]+)"):
        match = re.search(pattern, Path(filename).stem)
        if match:
            metadata["company_name"] = _clean_text(match.group(1))
            break

    round_match = re.search(r"\((\d+)차(?:수)?\)|\b(\d+)차수\b|과정\s*\((\d+)차\)", metadata["course_name"] or filename)
    if round_match:
        value = next((group for group in round_match.groups() if group), "")
        metadata["course_round"] = f"{value}차수" if value else ""

    metadata["course_name"] = re.sub(r"^\(([^)]+)\)\s*", "", metadata["course_name"])
    metadata["course_name"] = re.sub(r"\s*과정\s*\(\d+차\)\s*$", "", metadata["course_name"])
    return metadata



def _sheet_key_values(workbook, sheet_name: str) -> Dict[str, Any]:
    if sheet_name not in workbook.sheetnames:
        return {}
    worksheet = workbook[sheet_name]
    result: Dict[str, Any] = {}
    for row in worksheet.iter_rows(values_only=True):
        if len(row) < 2:
            continue
        key = _clean_text(row[0])
        if not key or key in {"항목", "입력 항목"} or key.startswith("▶"):
            continue
        result[key] = row[1]
    return result


def _parse_positive_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        parsed = int(float(value))
        return parsed if parsed > 0 else None
    except (TypeError, ValueError):
        match = re.search(r"\d+", str(value))
        return int(match.group()) if match else None


def _extract_question_settings(workbook) -> Dict[str, Any]:
    settings: Dict[str, Any] = {"common": {}, "instructors": {}, "subjective": {}}
    if "문항설정" not in workbook.sheetnames:
        return settings
    worksheet = workbook["문항설정"]
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        if len(row) < 4:
            continue
        try:
            number = int(row[0])
        except (TypeError, ValueError):
            continue
        kind = _clean_text(row[1])
        section = _clean_text(row[2])
        summary = _clean_text(row[3]).replace("\\n", "\n")
        slot_match = re.search(r"강사\s*(10|[1-9])\s*만족도\s*([12])", section)
        if slot_match:
            settings["instructors"][(int(slot_match.group(1)), int(slot_match.group(2)))] = {
                "section": section,
                "summary": summary,
                "kind": kind,
                "number": number,
            }
            continue
        if "좋았" in section:
            settings["subjective"]["subjective_good"] = {"section": section, "number": number}
        elif any(token in section for token in ("아쉬", "개선")):
            settings["subjective"]["subjective_bad"] = {"section": section, "number": number}
        elif number in {1, 2, 3, 4, 25}:
            category = {1: "overall", 2: "recommendation", 3: "schedule", 4: "difficulty", 25: "operation"}[number]
            settings["common"][category] = {"section": section, "summary": summary, "number": number}
    return settings


def _extract_template_config(workbook) -> Dict[str, Any]:
    basic = _sheet_key_values(workbook, "기본정보")
    # V2.3 simplified template keeps education overview fields in the same
    # 기본정보 sheet. Older five-sheet workbooks remain compatible because
    # values from 교육개요 are merged when that legacy sheet exists.
    overview = dict(basic)
    overview.update(_sheet_key_values(workbook, "교육개요"))
    instructor_slots: List[str] = [
        _clean_text(basic.get(f"강사{slot}", "")) for slot in range(1, 11)
    ]
    if not any(instructor_slots):
        legacy = _clean_text(basic.get("강사명", ""))
        if legacy:
            instructor_slots[0] = legacy
    has_slot_fields = any(re.fullmatch(r"강사(?:10|[1-9])", key) or key == "강사명" for key in basic)
    return {
        "basic": basic,
        "overview": overview,
        "instructor_slots": instructor_slots,
        "instructors": [name for name in instructor_slots if name],
        "use_instructor_slots": bool(has_slot_fields),
        "question_settings": _extract_question_settings(workbook),
    }


def _replace_instructor_tokens(text: str, instructor_slots: Sequence[str]) -> str:
    result = str(text or "")
    for slot in range(1, 11):
        name = instructor_slots[slot - 1] if slot <= len(instructor_slots) else ""
        result = result.replace(f"{{강사{slot}}}", name)
        result = re.sub(rf"강사\s*{slot}(?!\d)", name, result)
    result = re.sub(r"[ \t]+", " ", result)
    result = re.sub(r" *\n *", "\n", result)
    return result.strip()


def _assign_instructor_slots(questions: Sequence[ObjectiveQuestion]) -> None:
    named_slots: Dict[str, int] = {}
    next_slot = 1
    generic_index = 0
    for question in questions:
        if question.instructor_slot:
            next_slot = max(next_slot, question.instructor_slot + 1)
            continue
        if question.instructor_name:
            if question.instructor_name not in named_slots:
                named_slots[question.instructor_name] = next_slot
                next_slot += 1
            question.instructor_slot = named_slots[question.instructor_name]
        elif question.category == "instructor_generic":
            question.instructor_slot = generic_index // 2 + 1
            generic_index += 1


def _is_inactive_instructor_question(question_text: str, config: Dict[str, Any]) -> bool:
    """Return True for template placeholder questions whose instructor slot is blank.

    The manual-input template reserves rows for up to ten instructors. Rows for
    unused slots must never become choice/subjective questions, even when a
    workbook contains formulas, whitespace, or other placeholder values that
    make the row appear non-empty to the generic question classifier.
    """
    if not config.get("use_instructor_slots"):
        return False
    slot = _extract_instructor_slot(question_text)
    if not slot:
        return False
    instructor_slots: List[str] = list(config.get("instructor_slots", []))
    return slot > len(instructor_slots) or not _clean_text(instructor_slots[slot - 1])


def _filter_inactive_instructor_placeholders(
    objective_questions: List[ObjectiveQuestion],
    choice_questions: List[ChoiceQuestion],
    subjective_questions: List[SubjectiveQuestion],
    config: Dict[str, Any],
) -> Tuple[List[ObjectiveQuestion], List[ChoiceQuestion], List[SubjectiveQuestion]]:
    """Remove every parsed item belonging to an empty instructor slot."""
    objective_questions = [
        question for question in objective_questions
        if not _is_inactive_instructor_question(question.question, config)
    ]
    choice_questions = [
        question for question in choice_questions
        if not _is_inactive_instructor_question(question.question, config)
    ]
    subjective_questions = [
        question for question in subjective_questions
        if not _is_inactive_instructor_question(question.question, config)
    ]
    return objective_questions, choice_questions, subjective_questions


def _apply_template_config(
    objective_questions: List[ObjectiveQuestion],
    subjective_questions: List[SubjectiveQuestion],
    config: Dict[str, Any],
) -> Tuple[List[ObjectiveQuestion], List[SubjectiveQuestion], List[str]]:
    instructor_slots: List[str] = list(config.get("instructor_slots", []))
    instructors: List[str] = [name for name in instructor_slots if name]
    use_instructor_slots = bool(config.get("use_instructor_slots"))
    settings = config.get("question_settings", {})
    _assign_instructor_slots(objective_questions)

    filtered_objective: List[ObjectiveQuestion] = []
    for question in objective_questions:
        slot = int(question.instructor_slot or 0)
        if slot:
            # A configured workbook intentionally ignores empty instructor slots.
            if use_instructor_slots:
                display_name = instructor_slots[slot - 1] if slot <= len(instructor_slots) else ""
                if not display_name:
                    continue
                old_name = question.instructor_name
                question.question = _replace_instructor_tokens(question.question, instructor_slots)
                if old_name and old_name != display_name:
                    question.question = question.question.replace(old_name, display_name)
                question.instructor_name = display_name
                question.category = "instructor"
            metric_index = 2 if question.instructor_metric == "전문성" else 1
            setting = settings.get("instructors", {}).get((slot, metric_index))
            if setting:
                question.section_label = _replace_instructor_tokens(setting.get("section", ""), instructor_slots)
                question.summary_label = _replace_instructor_tokens(setting.get("summary", ""), instructor_slots)
        else:
            setting = settings.get("common", {}).get(question.category)
            if setting:
                question.section_label = setting.get("section") or question.section_label
                question.summary_label = setting.get("summary") or question.section_label
        filtered_objective.append(question)

    for question in subjective_questions:
        setting = settings.get("subjective", {}).get(question.category)
        if setting:
            question.section_label = setting.get("section") or question.section_label

    if not instructors:
        instructors = _dedupe_instructors(filtered_objective)
    return filtered_objective, subjective_questions, instructors


def _dedupe_instructors(questions: Sequence[ObjectiveQuestion]) -> List[str]:
    found: List[str] = []
    seen = set()
    for question in questions:
        name = question.instructor_name.strip()
        if name and name not in seen:
            seen.add(name)
            found.append(name)
    return found


def _build_insights(report: ReportData) -> List[str]:
    insights: List[str] = []
    scale_questions = report.objective_questions
    if scale_questions:
        normalized = [
            ((q.average - q.scale_min) / max(1, q.scale_max - q.scale_min), q)
            for q in scale_questions if q.valid_responses
        ]
        if normalized:
            _, top = max(normalized, key=lambda item: item[0])
            insights.append(f"가장 높은 평가는 '{top.section_label}'로 {top.average:.1f}점({top.scale_min}~{top.scale_max}점 척도)입니다.")
            _, low = min(normalized, key=lambda item: item[0])
            if low.id != top.id:
                insights.append(f"상대적으로 보완 여지가 큰 항목은 '{low.section_label}'로 {low.average:.1f}점입니다.")
    if len(report.instructors) >= 2:
        averages = []
        for name, questions in report.lecturers.items():
            if questions:
                averages.append((sum(q.average for q in questions) / len(questions), name))
        if averages:
            score, name = max(averages)
            insights.append(f"강사 비교에서는 {name} 강사의 평균 만족도가 {score:.1f}점으로 가장 높았습니다.")
    bad_questions = [q for q in report.subjective_questions if q.category == "subjective_bad" and q.answers]
    if bad_questions:
        insights.append(f"개선 의견은 총 {sum(len(q.answers) for q in bad_questions)}건이며, 원문 검토 후 운영 개선 과제로 반영할 수 있습니다.")
    return insights[:4]


def is_no_opinion(text: str) -> bool:
    core = re.sub(r"[\s.·,!?~]+", "", str(text or "")).lower()
    return core in NO_OPINION_CORES


def parse_excel(file_bytes: bytes, filename: str) -> ReportData:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    candidates: List[Dict[str, Any]] = []
    diagnostics: List[str] = []
    config = _extract_template_config(workbook)

    for worksheet in workbook.worksheets:
        if worksheet.title in {"입력가이드", "기본정보", "교육개요", "문항설정"}:
            continue
        candidate = _sheet_candidate(_sheet_matrix(worksheet), worksheet.title)
        if candidate:
            candidates.append(candidate)
            diagnostics.append(
                f"'{worksheet.title}' 시트: {candidate['orientation']}, 척도형 {len(candidate['objective'])}개, "
                f"선택형 {len(candidate['choices'])}개, 주관식 {len(candidate['subjective'])}개, "
                f"응답자 추정 {candidate['response_count']}명"
            )

    if not candidates:
        raise ValueError("설문응답 시트에서 문항과 응답 구조를 찾지 못했습니다. 첫 행에 문항, 아래 행에 응답이 있는지 확인해 주세요.")

    best = max(candidates, key=lambda item: item["quality"])
    metadata = _extract_metadata(_all_cell_texts(workbook), filename)
    objective_questions: List[ObjectiveQuestion] = best["objective"]
    choice_questions: List[ChoiceQuestion] = best["choices"]
    subjective_questions: List[SubjectiveQuestion] = best["subjective"]

    # The template includes placeholder rows for 강사1~강사10. Only slots with
    # a name in 기본정보 are active. Filter all parsed question types before
    # applying labels so unused rows cannot leak into 설문 구성 or 주관식 pages.
    objective_questions, choice_questions, subjective_questions = _filter_inactive_instructor_placeholders(
        objective_questions, choice_questions, subjective_questions, config
    )
    objective_questions, subjective_questions, instructors = _apply_template_config(
        objective_questions, subjective_questions, config
    )

    basic = config.get("basic", {})
    overview = config.get("overview", {})
    course_name = _clean_text(basic.get("과정명")) or metadata["course_name"]
    company_name = _clean_text(basic.get("고객사명")) or metadata["company_name"]
    schedule = _clean_text(basic.get("교육일정")) or metadata["schedule"]
    total_participants = _parse_positive_int(basic.get("수강인원"))
    cover_top_label = _clean_text(basic.get("표지_상단라벨")) or "결과보고서"
    cover_title1 = _clean_text(basic.get("표지_제목1")) or course_name
    cover_title2 = _clean_text(basic.get("표지_제목2")) or "결과보고서"

    report = ReportData(
        source_filename=filename,
        company_name=company_name,
        report_title=f"{course_name} 결과보고서" if course_name else "교육만족도 결과보고서",
        course_name=course_name,
        course_round=metadata["course_round"],
        schedule=schedule,
        delivery_method=_clean_text(overview.get("교육방식")),
        target_text=_clean_text(overview.get("교육대상")),
        total_participants=total_participants,
        response_count=best["response_count"],
        objective=_clean_text(overview.get("교육목표")),
        instructors=instructors,
        objective_questions=objective_questions,
        choice_questions=choice_questions,
        subjective_questions=subjective_questions,
        diagnostics=diagnostics,
        cover_top_label=cover_top_label,
        cover_title1=cover_title1,
        cover_title2=cover_title2,
    )
    report.insights = _build_insights(report)

    report.diagnostics.insert(0, f"분석에 사용한 시트: '{best['sheet_name']}'")
    if config.get("instructors"):
        report.diagnostics.append(f"기본정보에서 입력된 강사 {len(instructors)}명을 반영했습니다: {', '.join(instructors)}")
    scale_summary = Counter((q.scale_min, q.scale_max) for q in objective_questions)
    if scale_summary:
        report.diagnostics.append(
            "인식한 척도: " + ", ".join(f"{lo}~{hi}점 {count}문항" for (lo, hi), count in sorted(scale_summary.items()))
        )
    if not schedule:
        report.diagnostics.append("교육일정은 엑셀에서 찾지 못해 공란으로 두었습니다.")
    if not instructors:
        report.diagnostics.append("강사명은 기본정보와 설문 문항에서 찾지 못했습니다.")
    return report
